"""Date-ranged bot performance report → xlsx (BytesIO).

KPIs: SOP adherence (done/missed), completion + avg resolution, WhatsApp
adoption, escalations. ORM-based (runs in-process) so it replaces the old
API-scraping build_dashboard.py script for the dashboard export button.
"""
import io
from datetime import datetime, timezone, date, timedelta
from collections import defaultdict
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.task import Task, TaskStatus
from app.models.employee import Employee
from app.models.conversation import ConversationLog, Direction
from app.models.escalation import EscalationTicket
from app.models.sop import SOPExecution

# Same heuristic the old dashboard used, kept identical so numbers match.
DONE_KEYWORDS = {"done", "ho gaya", "hogya", "ho gya", "complete", "finished",
                 "kar diya", "kardiya", "the", "gaya", "ok", "yes", "haal"}
_FIRST_WORD_HINTS = {"done", "the", "gaya", "ok", "yes", "haal"}


def _classify_wa(text: str) -> bool:
    """True if a reply text reads as a WhatsApp 'done' confirmation."""
    t = (text or "").strip().lower()
    if not t:
        return False
    if t in DONE_KEYWORDS:
        return True
    first = t.split()
    return bool(first and first[0] in _FIRST_WORD_HINTS)


def _range(start: str | None, end: str | None):
    """Resolve (start_date, end_date) strings; default last 30 days incl today."""
    today = date.today()
    e = date.fromisoformat(end) if end else today
    s = date.fromisoformat(start) if start else e - timedelta(days=30)
    return s, e


def _hdr(ws, row, cols):
    fill = PatternFill("solid", fgColor="1F2937")
    for i, name in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="left")


def _widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report_xlsx(db, start: str | None = None, end: str | None = None) -> tuple[io.BytesIO, str]:
    s, e = _range(start, end)
    start_dt = datetime.combine(s, datetime.min.time(), tzinfo=timezone.utc)
    end_dt = datetime.combine(e, datetime.max.time(), tzinfo=timezone.utc)

    emps = {emp.id: emp for emp in db.execute(select(Employee)).scalars().all()}

    def dept(eid):
        emp = emps.get(eid)
        return emp.department if emp else "Unknown"

    def ename(eid):
        emp = emps.get(eid)
        return emp.name if emp else "Unknown"

    # Tasks assigned in range (the working set for this period).
    tasks = db.execute(
        select(Task).where(Task.assigned_at >= start_dt, Task.assigned_at <= end_dt)
    ).scalars().all()
    done = [t for t in tasks if t.status == TaskStatus.done]
    missed = [t for t in tasks if t.status == TaskStatus.missed]
    pending = [t for t in tasks if t.status in (TaskStatus.pending, TaskStatus.in_progress)]
    escalated = [t for t in tasks if t.status == TaskStatus.escalated]

    # WhatsApp adoption: which done tasks were closed by a 'done' reply.
    done_ids = [t.id for t in done]
    wa_ids = set()
    if done_ids:
        rows = db.execute(
            select(ConversationLog.task_id, ConversationLog.message_text).where(
                ConversationLog.task_id.in_(done_ids),
                ConversationLog.direction == Direction.inbound,
            )
        ).all()
        for tid, text in rows:
            if _classify_wa(text):
                wa_ids.add(tid)

    # Avg resolution (done tasks with both timestamps).
    durs = [
        (t.completed_at.replace(tzinfo=timezone.utc) - t.assigned_at.replace(tzinfo=timezone.utc)).total_seconds() / 3600
        for t in done if t.completed_at and t.assigned_at
    ]
    avg_res = round(sum(durs) / len(durs), 1) if durs else 0

    # SOP executions scheduled in range (done vs missed = adherence).
    execs = db.execute(
        select(SOPExecution).where(
            SOPExecution.scheduled_date >= s.isoformat(),
            SOPExecution.scheduled_date <= e.isoformat(),
        )
    ).scalars().all()

    messages = db.execute(
        select(ConversationLog).where(
            ConversationLog.created_at >= start_dt, ConversationLog.created_at <= end_dt
        )
    ).scalars().all()

    escs = db.execute(
        select(EscalationTicket).where(
            EscalationTicket.created_at >= start_dt, EscalationTicket.created_at <= end_dt
        )
    ).scalars().all()

    wb = Workbook()

    # ── Sheet 1: KPIs ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "KPIs"
    ws["A1"] = "Bot Performance Report"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = f"Period: {s.isoformat()} → {e.isoformat()}   |   Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    total = len(tasks)
    comp_rate = round(len(done) / total * 100, 1) if total else 0
    adherence = round(len(done) / (len(done) + len(missed)) * 100, 1) if (done or missed) else 0
    wa_pct = round(len(wa_ids) / len(done) * 100, 1) if done else 0
    kpis = [
        ("Total Tasks (assigned in period)", total),
        ("Completed", len(done)),
        ("Missed (SOP not done that day)", len(missed)),
        ("Pending / In-progress", len(pending)),
        ("Escalated", len(escalated)),
        ("Completion Rate", f"{comp_rate}%"),
        ("SOP Adherence (done / (done+missed))", f"{adherence}%"),
        ("Avg Resolution Time (hrs)", avg_res),
        ("WhatsApp Adoption (done via reply)", f"{wa_pct}%"),
        ("Employees on leave (now)", sum(1 for e2 in emps.values() if getattr(e2, "on_leave", False) and e2.is_active)),
        ("Messages exchanged", len(messages)),
        ("Escalations raised", len(escs)),
        ("Escalations open", sum(1 for x in escs if str(getattr(x, "status", "")) .endswith("open"))),
    ]
    _hdr(ws, 4, ["KPI", "Value"])
    for i, (k, v) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _widths(ws, [40, 18])

    # ── Sheet 2: SOP Adherence by department ────────────────────────────────
    ws = wb.create_sheet("SOP Adherence")
    agg = defaultdict(lambda: {"done": 0, "missed": 0, "leave": 0, "other": 0})
    for x in execs:
        d = dept(x.assigned_to_id)
        bucket = x.status if x.status in ("done", "missed", "leave") else "other"
        agg[d][bucket] += 1
    _hdr(ws, 1, ["Department", "Done", "Missed", "Leave", "Pending", "Total", "Adherence %"])
    for d, v in sorted(agg.items(), key=lambda kv: -(kv[1]["done"] + kv[1]["missed"])):
        tot = v["done"] + v["missed"] + v["leave"] + v["other"]
        # Leave days excluded from adherence — not the employee's fault.
        adh = round(v["done"] / (v["done"] + v["missed"]) * 100, 1) if (v["done"] + v["missed"]) else 0
        ws.append([d, v["done"], v["missed"], v["leave"], v["other"], tot, adh])
    _widths(ws, [28, 8, 8, 8, 9, 8, 12])

    # ── Sheet 3: Completion & Resolution by department ──────────────────────
    ws = wb.create_sheet("Completion")
    dagg = defaultdict(lambda: {"total": 0, "done": 0, "durs": []})
    for t in tasks:
        d = dept(t.assigned_to_id)
        dagg[d]["total"] += 1
        if t.status == TaskStatus.done:
            dagg[d]["done"] += 1
            if t.completed_at and t.assigned_at:
                dagg[d]["durs"].append(
                    (t.completed_at.replace(tzinfo=timezone.utc) - t.assigned_at.replace(tzinfo=timezone.utc)).total_seconds() / 3600
                )
    _hdr(ws, 1, ["Department", "Total", "Completed", "Completion %", "Avg Resolution (hrs)"])
    for d, v in sorted(dagg.items(), key=lambda kv: -kv[1]["total"]):
        rate = round(v["done"] / v["total"] * 100, 1) if v["total"] else 0
        avg = round(sum(v["durs"]) / len(v["durs"]), 1) if v["durs"] else 0
        ws.append([d, v["total"], v["done"], rate, avg])
    _widths(ws, [28, 8, 11, 13, 18])

    # ── Sheet 4: WhatsApp Adoption by employee ──────────────────────────────
    ws = wb.create_sheet("WhatsApp Adoption")
    eagg = defaultdict(lambda: {"wa": 0, "dash": 0})
    for t in done:
        if t.id in wa_ids:
            eagg[t.assigned_to_id]["wa"] += 1
        else:
            eagg[t.assigned_to_id]["dash"] += 1
    _hdr(ws, 1, ["Employee", "Department", "WhatsApp", "Dashboard", "Total", "WhatsApp %"])
    for eid, v in sorted(eagg.items(), key=lambda kv: -(kv[1]["wa"] + kv[1]["dash"])):
        tot = v["wa"] + v["dash"]
        pct = round(v["wa"] / tot * 100, 1) if tot else 0
        ws.append([ename(eid), dept(eid), v["wa"], v["dash"], tot, pct])
    _widths(ws, [22, 26, 10, 11, 8, 12])

    # ── Sheet 5: Escalations by department ──────────────────────────────────
    ws = wb.create_sheet("Escalations")
    escagg = defaultdict(lambda: {"open": 0, "resolved": 0, "total": 0})
    for x in escs:
        d = dept(x.employee_id)
        escagg[d]["total"] += 1
        key = "open" if str(getattr(x, "status", "")).endswith("open") else "resolved"
        escagg[d][key] += 1
    _hdr(ws, 1, ["Department", "Total", "Open", "Resolved"])
    for d, v in sorted(escagg.items(), key=lambda kv: -kv[1]["total"]):
        ws.append([d, v["total"], v["open"], v["resolved"]])
    _widths(ws, [28, 8, 8, 10])
    if not escs:
        ws.append(["(no escalations in period)", "", "", ""])

    # ── Sheet 6: All Tasks (every task, its stage, SOP completion time) ─────
    ws = wb.create_sheet("All Tasks")
    task_ids = [t.id for t in tasks]
    sop_ids = set()
    if task_ids:
        rows = db.execute(
            select(SOPExecution.task_id).where(SOPExecution.task_id.in_(task_ids))
        ).all()
        sop_ids = {r[0] for r in rows if r[0]}

    def fmt(dt):
        return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

    _hdr(ws, 1, ["Title", "Department", "Assignee", "Source", "Priority",
                 "Stage", "Assigned", "Due", "Completed (SOP)"])
    for t in sorted(tasks, key=lambda x: x.assigned_at or datetime.min.replace(tzinfo=timezone.utc)):
        is_sop = t.id in sop_ids
        ws.append([
            t.title,
            dept(t.assigned_to_id),
            ename(t.assigned_to_id),
            "SOP" if is_sop else "Manual",
            t.priority.value if hasattr(t.priority, "value") else str(t.priority),
            t.status.value if hasattr(t.status, "value") else str(t.status),
            fmt(t.assigned_at),
            fmt(t.due_date),
            fmt(t.completed_at) if is_sop else "",
        ])
    _widths(ws, [40, 22, 20, 9, 10, 12, 17, 17, 17])
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"bot_report_{s.isoformat()}_to_{e.isoformat()}.xlsx"
    return buf, fname


if __name__ == "__main__":
    # ponytail: self-check on the one non-trivial pure bit (WA classifier)
    assert _classify_wa("done") and _classify_wa("Done ✅") and _classify_wa("ho gaya")
    assert _classify_wa("ok bhai") and not _classify_wa("") and not _classify_wa("what is this")
    s, e = _range(None, None)
    assert (e - s).days == 30
    s2, e2 = _range("2026-06-01", "2026-06-10")
    assert s2.isoformat() == "2026-06-01" and e2.isoformat() == "2026-06-10"
    print("analytics_report self-check OK")
